
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.template import loader
from django.http import HttpResponse
from django import template

from django.contrib import messages
from django.contrib.auth.mixins import LoginRequiredMixin  # This is for authentication
from django.views import generic
from django.utils.safestring import mark_safe  # Explicitly mark a string as safe for (HTML) output
# purposes. The returned object can be used everywhere a string or unicode object is appropriate
from .utils import render_to_pdf, Calendar
from django.db.models import Q
from .utils import *
from .forms import *
from .models import *
import datetime
from datetime import date
from io import BytesIO
import xlsxwriter  ## This module is to wrinting files in the excel
import openpyxl  # Python library to read and write an excel file.
import calendar

from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash
from django.contrib.auth.forms import UserCreationForm, UserChangeForm, PasswordChangeForm
from .utils import render_to_pdf, Calendar
from .models import UserProfile
from django.views import generic
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect
from django.http import HttpResponse
from django.contrib import messages
from django.db.models import Q
import openpyxl
import calendar
from django.utils.safestring import mark_safe
from io import BytesIO
import xlsxwriter
from django.forms.models import model_to_dict
from django.contrib.auth.mixins import LoginRequiredMixin
from datetime import date
# model_to_dict(instance)
import datetime
from django.urls import reverse_lazy
from .models import *
from django.core.mail import send_mail
from datetime import datetime
from datetime import date,timedelta
from django.contrib.auth.models import User


class CalendarView(LoginRequiredMixin, generic.ListView):
    login_url = 'signup'
    model = Event
    template_name = 'accounts/calendar.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        d = get_date(self.request.GET.get('month', None))
        cal = Calendar(d.year, d.month)
        html_cal = cal.formatmonth(withyear=True)
        context['calendar'] = mark_safe(html_cal)
        context['prev_month'] = prev_month(d)
        context['next_month'] = next_month(d)
        return context


def get_date(req_day):
    if req_day:
        year, month = (int(x) for x in req_day.split('-'))
        return date(year, month, day=1)
    return datetime.today()


def prev_month(d):
    first = d.replace(day=1)
    prev_month = first - timedelta(days=1)
    month = 'month=' + str(prev_month.year) + '-' + str(prev_month.month)
    return month


def next_month(d):
    days_in_month = calendar.monthrange(d.year, d.month)[1]
    last = d.replace(day=days_in_month)
    next_month = last + timedelta(days=1)
    month = 'month=' + str(next_month.year) + '-' + str(next_month.month)
    return month

    '''
        context = super().get_context_data(**kwargs)
           d = get_date(self.request.GET.get('month', None))
           cal = Calendar(d.year, d.month)
           html_cal = cal.formatmonth(withyear=True)
           context['calendar'] = mark_safe(html_cal)
           context['prev_month'] = prev_month(d)
           context['next_month'] = next_month(d)
           return context
       '''


# def calendar_test(request):
#    msg     = None
#    success = False
#    context =""
#    #context = super().get_context_data(**kwargs)
#    d = get_date(request.GET.get('month', None))
#    cal = Calendar(d.year, d.month)
#    #print(cal)
#    html_cal = cal.formatmonth(withyear=True)
#    context = mark_safe(html_cal)
#    return render(request, "accounts/calendar.html", { "context" : context, "success" : success })


def create_event(request):
    print("test create event \n\n\n\n")
    form = EventForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        print(request.method)
        event = Event(
            user=request.user,
            title=request.POST['title'],
            description=request.POST['description'],
            start_date=request.POST['start_date'],
            end_date=request.POST['end_date'],
            start_time=request.POST['start_time'],
            end_time=request.POST['end_time']
        )
        event.save()
        messages.success(request, ('New Record Added Successfully...'))

        print("test create event ifififififififififi \n\n\n\n")
        title = request.POST['title']
        description = request.POST['description']
        start_date = request.POST['start_date']
        end_date = request.POST['end_date']
        start_time = request.POST['start_time']
        end_time = request.POST['end_time']
        # calendar_test(request)
        # return render(request, "accounts/calendar.html", {})

        return HttpResponseRedirect(reverse('calendar'))
    else:
        print("else part")

    args = Setup_details.objects.all()
    return render(request, 'accounts/event.html', {'form': form,'args':args})


class EventEdit(generic.UpdateView):
    model = Event
    fields = ['title', 'description', 'start_time', 'end_time']
    template_name = 'accounts/event.html'


# @login_required(login_url='signup')
def event_details(request, event_id):
    event = Event.objects.get(id=event_id)
    eventmember = EventMember.objects.filter(event=event)
    context = {
        'event': event,
        'eventmember': eventmember
    }
    return render(request, 'accounts/event_details.html', context)


def add_eventmember(request, event_id):
    forms = AddMemberForm()
    if request.method == 'POST':
        forms = AddMemberForm(request.POST)
        if forms.is_valid():
            member = EventMember.objects.filter(event=event_id)
            event = Event.objects.get(id=event_id)
            if member.count() <= 9:
                user = forms.cleaned_data['user']
                EventMember.objects.create(
                    event=event,
                    user=user
                )
                return redirect('calendar')
            else:
                print('--------------User limit exceed!-----------------')
    context = {
        'form': forms
    }
    return render(request, 'accounts/add_member.html', context)


class EventMemberDeleteView(generic.DeleteView):
    model = EventMember
    template_name = 'accounts/event_delete.html'
    success_url = reverse_lazy('calendar')


def asset_search(request):
    submitted = False
    if request.method == 'POST':
        print("home")
        if request.POST.get("export"):
            print("export")

        values = {}
        AssetNo = request.POST['AssetNo']
        Owner = request.POST['Owner']
        AssetTypeModel = request.POST['AssetTypeModel']
        Group = request.POST['Group']
        TeamName = request.POST['TeamName']
        Loc = request.POST['Loc']
        ProductLine = request.POST['ProductLine']
        Remark = request.POST['Remark']
        # store the data in session
        request.session['AssetNo'] = request.POST['AssetNo']
        request.session['Owner'] = request.POST['Owner']
        request.session['AssetTypeModel'] = request.POST['AssetTypeModel']
        request.session['Group'] = request.POST['Group']
        request.session['Loc'] = request.POST['Loc']
        request.session['TeamName'] = request.POST['TeamName']
        request.session['ProductLine'] = request.POST['ProductLine']
        request.session['Remark'] = request.POST['Remark']

        all_records = Asset_details.objects.filter(Q(Owner__contains=Owner) &
                                                   Q(AssetNo__contains=AssetNo) &
                                                   Q(Group__contains=Group) &
                                                   Q(TeamName__contains=TeamName) &
                                                   Q(AssetTypeModel__contains=AssetTypeModel) &
                                                   Q(Loc__contains=Loc) &
                                                   Q(ProductLine__contains=ProductLine) &
                                                   Q(Remark__contains=Remark)
                                                   )
        values = {'name': asset_search_display, 'all': all_records}
        return render(request, 'accounts/asset_search_display.html', values)
    else:
        data = Asset_details.objects.all()
        form = AssetDetailsForm()
        if 'submitted' in request.GET:
            submitted = True
    return render(request, 'accounts/asset_search.html', {'form': form, 'submitted': submitted, 'data': data})


def asset_search_display(request):
    all = Asset_details.objects.all()
    return render(request, 'accounts/asset_search_display.html', {'all': all})


def add_asset(request):
    values = {'form': AssetDetailsForm}
    form = AssetDetailsForm(request.POST)
    if form.is_valid():
        register = Asset_details(AssetNo=form.cleaned_data['AssetNo'],
                                 Owner=form.cleaned_data['Owner'],
                                 AssetTypeModel=form.cleaned_data['AssetTypeModel'],
                                 Group=form.cleaned_data['Group'],
                                 TeamName=form.cleaned_data['TeamName'],
                                 Loc=form.cleaned_data['Loc'],
                                 ProductLine=form.cleaned_data['ProductLine'],
                                 Remark=form.cleaned_data['Remark']
                                 )
        register.save()
        messages.success(request, 'New record added sucessfully.....')
    return render(request, 'accounts/add_asset.html', values)


def WriteToExcel(request, weather_data, town=None):
    AssetNo = request.session['AssetNo']
    Owner = request.session['Owner']
    AssetTypeModel = request.session['AssetTypeModel']
    Group = request.session['Group']
    TeamName = request.session['TeamName']
    Loc = request.session['Loc']
    ProductLine = request.session['ProductLine']
    Remark = request.session['Remark']
    all_data1 = Asset_details.objects.filter(Q(Owner__contains=Owner) &
                                             Q(AssetNo__contains=AssetNo) &
                                             Q(Group__contains=Group) &
                                             Q(TeamName__contains=TeamName) &
                                             Q(AssetTypeModel__contains=AssetTypeModel) &
                                             Q(Loc__contains=Loc) &
                                             Q(ProductLine__contains=ProductLine) &
                                             Q(Remark__contains=Remark)
                                             )
    output = BytesIO()
    workbook = xlsxwriter.Workbook(output)

    # Here we will adding the code to add data
    worksheet = workbook.add_worksheet()
    # Start from the first cell. Rows and columns are zero indexed.
    row = 0
    col = 0
    worksheet.write(row, col, "AssetNo")
    worksheet.write(row, col + 1, "Owner")
    worksheet.write(row, col + 2, "AssetTypeModel")
    worksheet.write(row, col + 3, "Group")
    worksheet.write(row, col + 4, "TeamName")
    worksheet.write(row, col + 5, "Loc")
    worksheet.write(row, col + 6, "ProductLine")
    worksheet.write(row, col + 7, "Remark")

    row = 1

    for data in all_data1:
        # print(data.AssetNo)
        worksheet.write(row, col, data.AssetNo)
        worksheet.write(row, col + 1, data.Owner)
        worksheet.write(row, col + 2, data.AssetTypeModel)
        worksheet.write(row, col + 3, data.Group)
        worksheet.write(row, col + 4, data.TeamName)
        worksheet.write(row, col + 5, data.Loc)
        worksheet.write(row, col + 6, data.ProductLine)
        worksheet.write(row, col + 7, data.Remark)
        row += 1
    workbook.close()
    xlsx_data = output.getvalue()
    # xlsx_data contains the Excel file
    return xlsx_data


def export_xls(request):
    print("welocme to export xls_files")
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment; filename=Report.xlsx'
    xlsx_data = WriteToExcel(request, "weather_period", "town")
    response.write(xlsx_data)
    return response


def export_pdf(request):
    AssetNo = request.session['AssetNo']
    Owner = request.session['Owner']
    AssetTypeModel = request.session['AssetTypeModel']
    Group = request.session['Group']
    TeamName = request.session['TeamName']
    Loc = request.session['Loc']
    ProductLine = request.session['ProductLine']
    Remark = request.session['Remark']
    all_data = Asset_details.objects.filter(Q(Owner__contains=Owner) &
                                            Q(AssetNo__contains=AssetNo) &
                                            Q(Group__contains=Group) &
                                            Q(TeamName__contains=TeamName) &
                                            Q(AssetTypeModel__contains=AssetTypeModel) &
                                            Q(Loc__contains=Loc) &
                                            Q(ProductLine__contains=ProductLine) &
                                            Q(Remark__contains=Remark)
                                            )

    report_name = AssetNo + "_" + Owner + "_" + AssetTypeModel + "_" + Group + "_" + TeamName + "_" + Loc + "_" + ProductLine

    data = {
        "report": report_name, "date": datetime.now(), "all": all_data,
    }
    pdf = render_to_pdf('accounts/pdf.html', data)
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename=Report.pdf'
    return response


def import_xls(request):
    print("upload xls")
    if "GET" == request.method:
        return render(request, 'accounts/import_excel.html', {})
    else:
        excel_file = request.FILES["excel_file"]
        # you may put validations here to check extension or file size
        wb = openpyxl.load_workbook(excel_file)
        # getting all sheets
        sheets = wb.sheetnames
        # getting a particular sheet
        worksheet = wb["Sheet1"]
        # getting active sheet
        active_sheet = wb.active
        # reading a cell
        excel_data = list()
        # iterating over the rows and
        # getting value from each cell in row
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))

            excel_data.append(row_data)
        data_valid = 0
        data = excel_data[0]
        if ("AssetNo" != data[0]):
            messages.success(request, ('AssetNo not macthed with data base field'))
            data_valid = 1
        if ("Owner" != data[1]):
            messages.success(request, ('Owner not macthed with data base field'))
            data_valid = 1
        if ("AssetTypeModel" != data[2]):
            messages.success(request, ('AssetTypeModel not macthed with data base field'))
            data_valid = 1
        if ("Group" != data[3]):
            messages.success(request, ('Group not macthed with data base field'))
            data_valid = 1
        if ("TeamName" != data[4]):
            messages.success(request, ('TeamName not macthed with data base field'))
            data_valid = 1
        if ("ProductLine" != data[5]):
            messages.success(request, ('ProductLine not macthed with data base field'))
            data_valid = all_data1 = Asset_details.objects.filter(Q(Owner__contains=Owner) &
                                                                  Q(AssetNo__contains=AssetNo) &
                                                                  Q(Group__contains=Group) &
                                                                  Q(TeamName__contains=TeamName) &
                                                                  Q(AssetTypeModel__contains=AssetTypeModel) &
                                                                  Q(ProductLine__contains=ProductLine) &
                                                                  Q(Remark__contains=Remark)
                                                                  )

        if (data_valid == 0):
            for data in excel_data[1:]:
                register = Asset_details(AssetNo=data[0],
                                         Owner=data[1],
                                         AssetTypeModel=data[2],
                                         Group=data[3],
                                         TeamName=data[4],
                                         ProductLine=data[5],
                                         Remark=data[6]
                                         )
                register.save()
        else:
            excel_data = []
            messages.success(request, ('import failed '))

        return render(request, 'accounts/import_excel.html', {"excel_data": excel_data})


def feedback(request):
    return render(request, 'feedback.html')


def table(request):
    assets = Asset_details.objects.all()
    return render(request, 'accounts/table.html', {'assets': assets})



@login_required(login_url="/login/")
def index(request):
    user=request.user
    args = Asset_details.objects.filter(Owner=user.username)
    return render(request, 'index.html', {'args': args})



@login_required(login_url="/login/")
def pages(request):
    context = {}
    # All resource paths end in .html.
    # Pick out the html file name from the url. And load that template.
    try:

        load_template = request.path.split('/')[-1]
        html_template = loader.get_template(load_template)
        return HttpResponse(html_template.render(context, request))

    except template.TemplateDoesNotExist:

        html_template = loader.get_template('page-404.html')
        return HttpResponse(html_template.render(context, request))

    except:

        html_template = loader.get_template('page-500.html')
        return HttpResponse(html_template.render(context, request))


def settings(request):
    return render(request, 'includes/settings.html')


def setup(request):
    values = {"form": setupDetailsForm}
    form = setupDetailsForm(request.POST)
    if form.is_valid():
        register = Setup_details(Host_name=form.cleaned_data['Host_name'],
                                 FQDN=form.cleaned_data['FQDN'],
                                 OS=form.cleaned_data['OS'],
                                 COM_port_details=form.cleaned_data['COM_port_details'],
                                 Other_details=form.cleaned_data['Other_details'],
                                 )
        register.save()
    messages.success(request, 'Setup details added succesfully.....')
    return render(request, 'accounts/setup.html', values)


def setup_display(request):
    setups = Setup_details.objects.all()
    return render(request, 'accounts/table1.html', {'setups': setups})


def setup_book(request):
    lists = Setup_details.objects.all()
    return render(request, 'accounts/book_event.html', {'lists': lists})


def delete_asset(request, id):
    id = int(id)
    all = Asset_details.objects.get(id=id)
    all.delete()
    return HttpResponseRedirect(reverse('asset_search_display'))


def edit_asset(request, id):
    all = Asset_details.objects.get(id=id)
    return render(request, 'accounts/edit_asset.html', {'all': all})


def delete_setup(request,id):
    id=int(id)
    setups = Setup_details.objects.get(id=id)
    setups.delete()
    return HttpResponseRedirect(reverse('setuptable'))

class EventMemberDeleteView(generic.DeleteView):
    model = EventMember
    template_name = 'accounts/event_delete.html'
    success_url = reverse_lazy('calendar')

def edit_setup(request, id):
    setups = Setup_details.objects.get(id=id)
    return render(request, 'accounts/edit_setup.html', {'setups': setups})



